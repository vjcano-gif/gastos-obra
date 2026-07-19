"""Importa el Excel historico de la constructora y hereda su clasificacion.

Se corre UNA vez (o cada vez que quieran resincronizar). Hace, en orden:

  1. Proyectos, con el %AIU del contrato (hoja LCLIENTES).
  2. Cortes por proyecto, con las fechas deducidas de los movimientos
     (su LCORTE no trae fechas, pero los movimientos si).
  3. Cruce contra las facturas ya cargadas desde Gmail: donde coincide,
     hereda proyecto / capitulo / actividad / corte SIN pisar nada.
  4. Carga como movimiento propio lo que nunca tuvo factura electronica
     (papel, cuenta de cobro, nomina): sin eso el costo no cuadra.

Es idempotente: se puede volver a correr. Lo que ya existe se actualiza,
y los movimientos importados se reconocen por `origen_matriz`.

Uso:
    python -u -m worker.importar_matriz ruta/al/archivo.xlsx
    python -u -m worker.importar_matriz ruta.xlsx --simular   (no escribe)
"""
from __future__ import annotations

import sys

import openpyxl

from . import matriz
from .config import Config
from .paginacion import traer_todo
from .storage import Store


def _todas(sb, tabla: str, uid: str, columnas: str = "*") -> list[dict]:
    """Tabla completa del workspace, paginando el tope de PostgREST."""
    return traer_todo(
        sb.table(tabla).select(columnas).eq("user_id", uid).order("id")
    )


class Importador:
    """Recibe un cliente de Supabase ya construido, no la configuracion.

    Asi lo pueden usar los dos caminos con el MISMO codigo: la linea de
    comandos (con service_role) y la pantalla de la app (con la sesion del
    usuario, donde el RLS sigue aplicando). Duplicar esta logica en la app
    habria sido garantizar que las dos versiones se separen con el tiempo.

    `ruta` puede ser una ruta en disco o cualquier objeto tipo archivo
    (lo que devuelve el file_uploader de Streamlit).
    """

    def __init__(self, sb, uid: str, ruta, simular: bool = False):
        self.sb = sb
        self.uid = uid
        self.ruta = ruta
        self.simular = simular
        self.resumen: dict[str, int] = {}

    @classmethod
    def desde_config(cls, cfg: Config, ruta: str, simular: bool = False) -> "Importador":
        return cls(Store(cfg).sb, cfg.user_id, ruta, simular)

    def _contar(self, clave: str, n: int = 1) -> None:
        self.resumen[clave] = self.resumen.get(clave, 0) + n

    # ------------------------------------------------------------ lectura
    def leer(self) -> None:
        wb = openpyxl.load_workbook(self.ruta, read_only=True, data_only=True)
        self.gastos = matriz.leer_gastos(wb["MATRIZ GASTOS"].iter_rows(values_only=True))
        self.ingresos = matriz.leer_ingresos(wb["MATRIZ INGRESOS"].iter_rows(values_only=True))
        self.proyectos_excel = matriz.leer_proyectos(wb["LCLIENTES"].iter_rows(values_only=True))
        wb.close()
        print(f"Leidos: {len(self.gastos)} gastos, {len(self.ingresos)} ingresos, "
              f"{len(self.proyectos_excel)} proyectos")

    # ---------------------------------------------------------- proyectos
    def sincronizar_proyectos(self) -> None:
        actuales = {
            matriz.norm(p["nombre"]): p
            for p in _todas(self.sb, "proyectos", self.uid)
        }
        self.ids_proyecto = {}
        for p in self.proyectos_excel:
            clave = matriz.norm(p["nombre"])
            existente = actuales.get(clave)
            fila = {
                "nombre": p["nombre"],
                "pct_aiu": p["pct_aiu"],
                "estado": p["estado"],
            }
            if existente:
                # El %AIU del contrato manda; lo demas no se toca para no
                # pisar lo que hayan ajustado en la app.
                if not self.simular:
                    self.sb.table("proyectos").update(fila).eq("id", existente["id"]).execute()
                self.ids_proyecto[clave] = existente["id"]
                self._contar("proyectos_actualizados")
            else:
                fila |= {"user_id": self.uid, "codigo": _codigo_proyecto(p["nombre"])}
                if self.simular:
                    self.ids_proyecto[clave] = f"simulado:{clave}"
                else:
                    r = self.sb.table("proyectos").insert(fila).execute()
                    self.ids_proyecto[clave] = r.data[0]["id"]
                self._contar("proyectos_nuevos")

    # ------------------------------------------------------------- cortes
    def sincronizar_cortes(self) -> None:
        rangos = matriz.rangos_de_cortes(self.gastos)
        actuales = {}
        for c in _todas(self.sb, "cortes", self.uid):
            actuales[(c["proyecto_id"], matriz.norm(c["nombre"]))] = c["id"]

        self.ids_corte = {}
        for (proyecto, corte), r in sorted(
            rangos.items(), key=lambda kv: (kv[0][0], matriz.numero_de_corte(kv[0][1]))
        ):
            proyecto_id = self.ids_proyecto.get(proyecto)
            if not proyecto_id:
                self._contar("cortes_sin_proyecto")
                continue
            numero = matriz.numero_de_corte(corte)
            fila = {
                "proyecto_id": proyecto_id,
                "numero": numero,
                "nombre": f"Corte {numero}",
                "fecha_inicio": r["fecha_inicio"].isoformat(),
                "fecha_fin": r["fecha_fin"].isoformat(),
                "descripcion": f"Deducido de {r['movimientos']} movimientos del Excel",
            }
            existente = actuales.get((proyecto_id, matriz.norm(fila["nombre"])))
            if existente:
                if not self.simular:
                    self.sb.table("cortes").update(fila).eq("id", existente).execute()
                self.ids_corte[(proyecto, corte)] = existente
                self._contar("cortes_actualizados")
            else:
                if self.simular:
                    self.ids_corte[(proyecto, corte)] = f"simulado:{proyecto}:{corte}"
                else:
                    res = self.sb.table("cortes").insert({"user_id": self.uid, **fila}).execute()
                    self.ids_corte[(proyecto, corte)] = res.data[0]["id"]
                self._contar("cortes_nuevos")

    # -------------------------------------------------------- dimensiones
    def cargar_dimensiones(self) -> None:
        self.ids_capitulo = {
            str(c["codigo"]): c["id"]
            for c in _todas(self.sb, "capitulos", self.uid)
            if c.get("codigo")
        }
        self.ids_actividad = {
            str(a["codigo"]): a["id"]
            for a in _todas(self.sb, "actividades", self.uid)
            if a.get("codigo")
        }
        if not self.ids_capitulo:
            print("  AVISO: no hay capitulos con codigo. Corre primero la siembra "
                  "del catalogo desde Configuracion, o no se heredara la clasificacion.")

    # -------------------------------------------------------------- cruce
    def cruzar(self) -> None:
        facturas = _todas(self.sb, "facturas", self.uid)
        print(f"Facturas ya cargadas: {len(facturas)}")
        idx = matriz.indexar_facturas(facturas)
        ids = {
            "proyectos": self.ids_proyecto,
            "capitulos": self.ids_capitulo,
            "actividades": self.ids_actividad,
            "cortes": self.ids_corte,
        }

        self.no_cruzadas = []
        for fila in self.gastos:
            factura, motivo = matriz.emparejar(fila, idx)
            if factura is None:
                self._contar(f"no_cruzo_{motivo}")
                self.no_cruzadas.append(fila)
                continue
            self._contar(f"cruzo_{motivo}")
            cambios = matriz.cambios_heredables(fila, factura, ids)
            if not cambios:
                # OJO con el nombre: no puede empezar por "cruzo_" porque
                # cruce() suma ese prefijo para contar los emparejamientos.
                # Cuando se llamaba "cruzo_sin_cambios" se contaba dos veces
                # y el informe llegaba a decir "2 de 1 (200%)".
                self._contar("ya_estaba_completa")
                continue
            if not self.simular:
                self.sb.table("facturas").update(cambios).eq("id", factura["id"]).execute()
            self._contar("facturas_enriquecidas")
            for campo in cambios:
                self._contar(f"campo_{campo}")

    # ------------------------------------- movimientos que no tienen factura
    def cargar_no_cruzadas(self) -> None:
        """Carga como costo propio lo que nunca tuvo factura electronica.

        Solo entran las filas con proyecto y valor: sin eso no aportan al
        costo y solo ensuciarian la revision.
        """
        existentes = {
            f.get("origen_matriz")
            for f in _todas(self.sb, "facturas", self.uid, "id,origen_matriz")
            if f.get("origen_matriz")
        }
        nuevas = []
        for fila in self.no_cruzadas:
            origen = f"MATRIZ GASTOS!{fila['fila_excel']}"
            if origen in existentes:
                self._contar("importadas_ya_existian")
                continue
            proyecto_id = self.ids_proyecto.get(matriz.norm(fila["proyecto"]))
            if not proyecto_id or not fila["subtotal"]:
                self._contar("descartadas_sin_proyecto_o_valor")
                continue
            nuevas.append(
                {
                    "user_id": self.uid,
                    "proyecto_id": proyecto_id,
                    "capitulo_id": self.ids_capitulo.get(fila.get("capitulo_codigo")),
                    "actividad_id": self.ids_actividad.get(fila.get("actividad_codigo")),
                    "corte_id": self.ids_corte.get(
                        (matriz.norm(fila["proyecto"]), fila.get("corte"))
                    ),
                    "tipo_documento": _tipo_documento(fila.get("documento")),
                    "numero": fila.get("numero"),
                    "proveedor_nombre": fila.get("proveedor"),
                    "proveedor_nit": fila.get("nit"),
                    "fecha_emision": fila["fecha"].isoformat() if fila.get("fecha") else None,
                    "descripcion": fila.get("descripcion"),
                    "concepto": fila.get("descripcion"),
                    "iva": fila.get("iva") or 0,
                    "total": fila["subtotal"],
                    "retenciones_xml": fila.get("retenciones") or 0,
                    "forma_pago": fila.get("forma_pago"),
                    "metodo_pago": fila.get("metodo_pago"),
                    "pagador": fila.get("pagador"),
                    "legalizacion": fila.get("legalizacion"),
                    "exento_aiu": fila.get("exento_aiu", False),
                    "estado_pago": fila.get("estado_pago", "pendiente"),
                    # Viene del Excel, no de un XML validado: entra con
                    # confianza baja y en estado 'asignada' (clasificada
                    # pero sin aprobar), nunca como aprobada.
                    "fuente": "matriz",
                    "confianza": "baja",
                    "estado": "anulada" if fila.get("estado_pago") == "anulada" else "asignada",
                    "origen_matriz": origen,
                }
            )
        self._contar("importadas_nuevas", len(nuevas))
        if not self.simular:
            for i in range(0, len(nuevas), 200):     # lotes: el insert masivo falla
                self.sb.table("facturas").insert(nuevas[i:i + 200]).execute()

    # ---------------------------------------------------------- anticipos
    def cargar_anticipos(self) -> None:
        existentes = {
            a.get("origen_matriz")
            for a in _todas(self.sb, "anticipos", self.uid, "id,origen_matriz")
            if a.get("origen_matriz")
        }
        nuevos = []
        for fila in self.ingresos:
            origen = f"MATRIZ INGRESOS!{fila['fila_excel']}"
            proyecto_id = self.ids_proyecto.get(matriz.norm(fila["proyecto"]))
            if origen in existentes or not proyecto_id or not fila["valor"]:
                self._contar("anticipos_omitidos")
                continue
            nuevos.append(
                {
                    "user_id": self.uid,
                    "proyecto_id": proyecto_id,
                    "corte_id": self.ids_corte.get(
                        (matriz.norm(fila["proyecto"]), fila.get("corte"))
                    ),
                    "fecha": fila["fecha"].isoformat() if fila.get("fecha") else None,
                    "valor": fila["valor"],
                    "modo_pago": fila["modo_pago"],
                    "detalle": fila.get("detalle"),
                    "legalizacion": fila.get("legalizacion"),
                    "origen_matriz": origen,
                }
            )
        nuevos = [n for n in nuevos if n["fecha"]]
        self._contar("anticipos_nuevos", len(nuevos))
        if not self.simular:
            for i in range(0, len(nuevos), 200):
                self.sb.table("anticipos").insert(nuevos[i:i + 200]).execute()

    # ------------------------------------------------------------ informe
    def cruce(self) -> tuple[int, int]:
        """(filas que cruzaron, filas leidas). Lo usan el informe de
        consola y la pantalla de la app, para no calcularlo dos veces."""
        cruzadas = sum(v for k, v in self.resumen.items() if k.startswith("cruzo_"))
        return cruzadas, len(getattr(self, "gastos", []))

    def informar(self) -> None:
        print("\n" + "=" * 60)
        print("RESUMEN" + ("  (SIMULACION: no se escribio nada)" if self.simular else ""))
        print("=" * 60)
        for k in sorted(self.resumen):
            print(f"  {k:<40} {self.resumen[k]:>7}")
        cruzadas, total = self.cruce()
        total = total or 1
        print(f"\n  Cruce: {cruzadas} de {len(self.gastos)} ({cruzadas * 100 // total}%)")

    def correr(self) -> None:
        self.leer()
        self.sincronizar_proyectos()
        self.sincronizar_cortes()
        self.cargar_dimensiones()
        self.cruzar()
        self.cargar_no_cruzadas()
        self.cargar_anticipos()
        self.informar()


def _codigo_proyecto(nombre: str) -> str:
    """Codigo corto para renombrar archivos: "Casa Vieja 61" -> CASAVIEJA61."""
    limpio = "".join(c for c in matriz.norm(nombre).upper() if c.isalnum())
    return limpio[:20] or "PROYECTO"


def _tipo_documento(documento: str | None) -> str:
    d = matriz.norm(documento)
    if "nota credito" in d:
        return "nota_credito"
    if "nota debito" in d:
        return "nota_debito"
    if "cuenta de cobro" in d:
        return "cuenta_cobro"
    if "anticipo" in d or "recibo de caja" in d:
        return "consignacion"
    if "factura" in d or "documento equivalente" in d:
        return "factura"
    return "otro"


def main() -> None:
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        raise SystemExit("Uso: python -m worker.importar_matriz <archivo.xlsx> [--simular]")
    cfg = Config()
    if not (cfg.supabase_url and cfg.supabase_service_key and cfg.user_id):
        raise SystemExit("Faltan SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY o APP_USER_ID")
    Importador.desde_config(cfg, args[0], simular="--simular" in sys.argv).correr()


if __name__ == "__main__":
    main()
