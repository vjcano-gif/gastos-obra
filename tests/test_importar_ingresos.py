"""Parseo de la MATRIZ DE INGRESOS (Excel) -> filas para la tabla `anticipos`.

El importador es la vía por la que la constructora sube su histórico de
abonos del cliente. Si el parseo cambia (columnas, formato de valor, modo de
pago), Arrayanes 40 volvería a mostrar ingresos en cero — el bug que este
módulo resuelve.
"""
import sys
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "app"))

from lib import importar_ingresos  # noqa: E402

_ENC = ["Fecha", "Proyecto", "Corte", "Detalle", "Total", "Modo de Pago", "Encima / Debajo"]


def _excel(filas, encabezados=None):
    wb = Workbook()
    ws = wb.active
    ws.append(encabezados or _ENC)
    for f in filas:
        ws.append(f)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parsea_columnas_y_normaliza_modo_y_legalizacion():
    contenido = _excel([
        ["2024-05-07", "Arrayanes 40", "Corte 1", "Transferencia", 80000000, "Transferencia", "Encima"],
        ["7/05/2024", "Arrayanes 40", "Corte 3", "Transferencia", 59000000, "Transferencia", "Encima"],
    ])
    d = importar_ingresos.parsear_excel(contenido)
    assert len(d) == 2
    assert d.iloc[0]["proyecto"] == "Arrayanes 40"
    assert d.iloc[0]["valor"] == 80000000
    assert d.iloc[0]["modo_pago"] == "bancos"        # transferencia -> bancos
    assert d.iloc[0]["legalizacion"] == "encima"
    assert d.iloc[0]["corte"] == "Corte 1"
    assert d.iloc[0]["fecha"] == "2024-05-07"
    # "7/05/2024" (día primero, como en su Excel) -> 7 de mayo, no 5 de julio
    assert d.iloc[1]["fecha"] == "2024-05-07"


def test_descarta_filas_sin_proyecto_o_sin_valor():
    contenido = _excel([
        ["2024-05-07", "", "", "x", 1000, "", ""],                    # sin proyecto
        ["2024-05-07", "Casa Vieja 47", "", "Zular", 0, "", ""],      # valor 0
        ["2024-05-07", "Casa Vieja 47", "", "Zular", 80000000, "", "Encima"],  # válida
    ])
    d = importar_ingresos.parsear_excel(contenido)
    assert len(d) == 1
    assert d.iloc[0]["proyecto"] == "Casa Vieja 47"
    assert d.iloc[0]["modo_pago"] == "por_identificar"   # modo vacío
    assert d.iloc[0]["corte"] is None
    assert d.iloc[0]["legalizacion"] == "encima"


def test_modos_de_pago_variados():
    assert importar_ingresos.modo_pago_slug("Transferencia") == "bancos"
    assert importar_ingresos.modo_pago_slug("Consignación") == "bancos"
    assert importar_ingresos.modo_pago_slug("Efectivo") == "efectivo"
    assert importar_ingresos.modo_pago_slug("Pago Directo") == "pago_directo"
    assert importar_ingresos.modo_pago_slug("") == "por_identificar"
    assert importar_ingresos.modo_pago_slug("otra cosa") == "por_identificar"


def test_valor_como_texto_con_formato_colombiano():
    contenido = _excel([
        ["2024-05-07", "Arrayanes 40", "", "x", "$ 103.311.060", "Transferencia", "Encima"],
    ])
    d = importar_ingresos.parsear_excel(contenido)
    assert d.iloc[0]["valor"] == 103311060.0


def test_elige_la_hoja_matriz_ingresos_entre_varias():
    """El libro real trae ~18 hojas; la de ingresos no es la primera y tiene
    columnas de sobra (Día, Mes, Año) que deben ignorarse."""
    from datetime import datetime
    wb = Workbook()
    wb.active.title = "LCORTE"          # hoja decoy, primera
    wb.active.append(["algo", "otra"])
    ws = wb.create_sheet("MATRIZ INGRESOS")
    ws.append(["Fecha", "Día", "Mes", "Año", "Proyecto", "Corte", "Detalle",
               "Total", "Modo de Pago", "Encima / Debajo"])
    ws.append([datetime(2024, 5, 7), 7, "May", 2024, "Arrayanes 40", "Corte 1",
               "Transferencia", 80000000, "Transferencia", "Encima"])
    buf = BytesIO()
    wb.save(buf)
    d = importar_ingresos.parsear_excel(buf.getvalue())
    assert len(d) == 1
    assert d.iloc[0]["proyecto"] == "Arrayanes 40"
    assert d.iloc[0]["valor"] == 80000000
    assert d.iloc[0]["fecha"] == "2024-05-07"       # fecha real de Excel
    assert d.iloc[0]["modo_pago"] == "bancos"
    assert list(d.columns) == importar_ingresos.COLUMNAS   # sin Día/Mes/Año


def test_celdas_vacias_son_none_no_nan():
    """Un NaN de pandas no es serializable a JSON y reventaba el insert a
    Supabase. Las celdas vacías deben salir como None."""
    import math
    contenido = _excel([
        ["2025-01-24", "Smart Fit", "Sin Corte", None, 2552000, None, "Encima"],
    ])
    d = importar_ingresos.parsear_excel(contenido)
    fila = d.iloc[0]
    assert fila["detalle"] is None
    assert fila["modo_pago"] == "por_identificar"
    assert not any(isinstance(v, float) and math.isnan(v) for v in fila.tolist())


def test_archivo_sin_columnas_esperadas_avisa():
    wb = Workbook()
    ws = wb.active
    ws.append(["algo", "otra"])
    ws.append([1, 2])
    buf = BytesIO()
    wb.save(buf)
    try:
        importar_ingresos.parsear_excel(buf.getvalue())
        raise AssertionError("debió lanzar ValueError por columnas faltantes")
    except ValueError:
        pass


if __name__ == "__main__":
    for nombre, fn in sorted(globals().items()):
        if nombre.startswith("test_"):
            fn()
            print("OK  ", nombre)
