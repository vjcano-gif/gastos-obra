"""Las plantillas descargables deben poder subirse SIN NOVEDAD: se prueba que
cada una la lea su propio importador y que sigan en sync con él."""
import io
import sys
from pathlib import Path

from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(RAIZ / "app"))
sys.path.insert(0, str(RAIZ))

from lib import importar_ingresos, plantillas  # noqa: E402


def test_plantillas_son_xlsx_validos():
    for b in (plantillas.matriz_ingresos(), plantillas.matriz_gastos()):
        assert isinstance(b, bytes)
        assert b[:2] == b"PK"          # firma zip de un .xlsx
        assert len(b) > 500


def test_plantilla_ingresos_la_lee_el_parser_y_omite_el_ejemplo():
    """La plantilla la lee el MISMO parser sin error, y su fila de ejemplo
    (marcada EJEMPLO) NO se importa: subirla tal cual no mete datos falsos."""
    d = importar_ingresos.parsear_excel(plantillas.matriz_ingresos())
    assert d.empty                                     # el ejemplo se ignora
    assert list(d.columns) == importar_ingresos.COLUMNAS


def test_encabezados_ingresos_los_reconoce_el_parser():
    for h in plantillas.ENCABEZADOS_INGRESOS:
        assert importar_ingresos._norm(h) in importar_ingresos._ENCABEZADOS


# El nombre legible de cada columna de la plantilla, atado al CAMPO que lee el
# importador en esa posición. Si alguien reordena worker.matriz.COL, este mapa
# deja de cuadrar y el test falla (no basta con que la posición exista).
_LABEL_A_CAMPO = {
    "Proyecto": "proyecto", "ID Capítulo": "id_capitulo", "Capítulo": "capitulo",
    "Corte": "corte", "ID Actividad": "id_actividad", "Actividad": "actividad",
    "Fecha": "fecha", "Proveedor": "proveedor", "NIT": "nit", "Documento": "documento",
    "Número": "numero", "Descripción": "descripcion", "Valor bruto": "valor_bruto",
    "Descuento": "descuento", "IVA": "iva", "Subtotal": "subtotal",
    "Retenciones": "retenciones", "Total a pagar": "total_a_pagar",
    "Forma de pago": "forma_pago", "Estado": "estado", "Medio de pago": "medio_pago",
    "Pagador": "pagador", "Legalización": "legalizacion",
    "Fecha vencimiento": "fecha_vencimiento", "Concepto": "concepto",
    "Fecha de pago": "fecha_pago", "Valor pagado": "valor_pagado",
    "Valor pagado 2": "valor_pagado2", "Saldo calculado": "saldo",
    "Exento AIU": "exento_aiu", "% AIU": "pct_aiu", "Comisión": "comision",
}


def test_plantilla_gastos_en_sync_con_las_posiciones_del_importador():
    """Cada columna de la plantilla debe caer en la MISMA posición que el campo
    que el importador lee ahí (detecta reordenamientos de COL, no solo altas/bajas)."""
    from worker.matriz import COL
    for pos, label in plantillas._COLS_GASTOS.items():
        campo = _LABEL_A_CAMPO[label]
        assert COL[campo] == pos, f"'{label}' está en la col {pos} pero el importador lee '{campo}' en la col {COL[campo]}"


def test_plantilla_gastos_respeta_el_orden_por_posicion():
    ws = load_workbook(io.BytesIO(plantillas.matriz_gastos()))["MATRIZ GASTOS"]
    assert ws.cell(row=1, column=1).value == "Proyecto"
    assert ws.cell(row=1, column=31).value == "Estado"
    assert ws.cell(row=1, column=49).value == "Comisión"


if __name__ == "__main__":
    for nombre, fn in sorted(globals().items()):
        if nombre.startswith("test_"):
            fn()
            print("OK  ", nombre)
