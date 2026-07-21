"""Plantillas .xlsx descargables para los cargadores de la app.

Genera en memoria el archivo con los encabezados EXACTOS que espera cada
importador, una fila de ejemplo y una hoja de Instrucciones, para que quien
sube el archivo no falle por formato. Es la contraparte de los parsers
(lib.importar_ingresos y worker.matriz): si un importador cambia sus columnas,
su plantilla debe cambiar aquí también — el test lo verifica.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_NARANJA = "D85A30"
_GRIS = "EEEEEE"


def _titulos(ws, headers, ancho_min=12):
    fill = PatternFill("solid", fgColor=_NARANJA)
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
        if h:
            ws.column_dimensions[c.column_letter].width = max(ancho_min, len(str(h)) + 3)


def _hoja_instrucciones(wb, filas):
    ins = wb.create_sheet("Instrucciones")
    ins.append(["Columna", "Qué va", "Ejemplo / valores válidos"])
    for c in ins[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=_NARANJA)
    for f in filas:
        ins.append(f)
    ins.column_dimensions["A"].width = 22
    ins.column_dimensions["B"].width = 48
    ins.column_dimensions["C"].width = 52
    for fila in ins.iter_rows(min_row=2):
        for c in fila:
            c.alignment = Alignment(vertical="top", wrap_text=True)


# --------------------------------------------------------------- ingresos
ENCABEZADOS_INGRESOS = ["Fecha", "Proyecto", "Corte", "Detalle", "Total",
                        "Modo de Pago", "Encima / Debajo"]


def matriz_ingresos() -> bytes:
    """Plantilla de la MATRIZ DE INGRESOS (abonos del cliente)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MATRIZ INGRESOS"
    _titulos(ws, ENCABEZADOS_INGRESOS, ancho_min=16)
    ws.append(["2025-01-24", "Arrayanes 40", "Corte 1", "Transferencia RC 71",
               80000000, "Transferencia", "Encima"])
    ws.append(["2025-02-13", "Casa Vieja 47", "Sin Corte", "Consignación Zular",
               45000000, "Consignación", "Debajo"])
    _hoja_instrucciones(wb, [
        ["Fecha", "Fecha del abono del cliente.", "2025-01-24 o 24/01/2025"],
        ["Proyecto", "Nombre EXACTO del proyecto, tal como está creado en la app.", "Arrayanes 40"],
        ["Corte", "Corte de obra al que aplica; vacío o 'Sin Corte' si no aplica.", "Corte 1"],
        ["Detalle", "Descripción o número de recibo (RC).", "Transferencia RC 71"],
        ["Total", "Valor en pesos, sin puntos ni signo $.", "80000000"],
        ["Modo de Pago", "Cómo entró la plata.", "Transferencia / Consignación / Efectivo / Pago Directo"],
        ["Encima / Debajo", "Si el abono va por encima o por debajo del presupuesto.", "Encima / Debajo"],
        ["", "Nota: la app empareja Proyecto y Corte por nombre y NO duplica lo ya cargado.", ""],
    ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------------------------------------ presupuesto
ENCABEZADOS_PRESUPUESTO = ["Capítulo", "Actividad", "Subactividad", "Unidad",
                          "Cantidad", "Costo unitario", "Costo total"]


def presupuesto() -> bytes:
    """Plantilla del PRESUPUESTO por actividad (flujo semanal)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "PRESUPUESTO"
    _titulos(ws, ENCABEZADOS_PRESUPUESTO, ancho_min=16)
    ws.append(["EXCAVACIONES, FUNDACIONES Y CONCRETOS", "Fundaciones",
               "Vaciado de zapatas", "m3", 120, 450000, 54000000])
    ws.append(["MAMPOSTERIA", "", "Muros en bloque", "m2", 300, 38000, 11400000])
    _hoja_instrucciones(wb, [
        ["Capítulo", "Nombre del capítulo, tal como está en la app.", "MAMPOSTERIA"],
        ["Actividad", "Actividad del capítulo (opcional).", "Fundaciones"],
        ["Subactividad", "Detalle libre de la línea (opcional).", "Vaciado de zapatas"],
        ["Unidad", "Unidad de medida.", "m3 / m2 / gl / uds"],
        ["Cantidad", "Cantidad presupuestada.", "120"],
        ["Costo unitario", "Valor por unidad, en pesos sin puntos.", "450000"],
        ["Costo total", "Valor total; si se deja vacío se calcula cantidad × unitario.", "54000000"],
        ["", "La app empareja Capítulo y Actividad por nombre. Debe existir al "
         "menos uno de: capítulo, actividad o subactividad, y un valor.", ""],
    ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ----------------------------------------------------------------- gastos
# La hoja "MATRIZ GASTOS" se lee por POSICIÓN de columna (los encabezados del
# archivo real traen espacios y tildes inconsistentes). Este es el nombre
# legible de cada posición 1..49 que usa worker/matriz.py; las que no se leen
# van en blanco. NO se deben mover de lugar.
_COLS_GASTOS = {
    1: "Proyecto", 2: "ID Capítulo", 3: "Capítulo", 4: "Corte",
    5: "ID Actividad", 6: "Actividad", 7: "Fecha", 12: "Proveedor",
    13: "NIT", 15: "Documento", 16: "Número", 17: "Descripción",
    18: "Valor bruto", 19: "Descuento", 20: "IVA", 26: "Subtotal",
    27: "Retenciones", 29: "Total a pagar", 30: "Forma de pago",
    31: "Estado", 32: "Medio de pago", 33: "Pagador", 34: "Legalización",
    36: "Fecha vencimiento", 37: "Concepto", 41: "Fecha de pago",
    42: "Valor pagado", 44: "Valor pagado 2", 46: "Saldo calculado",
    47: "Exento AIU", 48: "% AIU", 49: "Comisión",
}


def matriz_gastos() -> bytes:
    """Plantilla de la MATRIZ GASTOS (movimientos contables), respetando las
    posiciones de columna que lee el importador."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MATRIZ GASTOS"
    n = max(_COLS_GASTOS)
    headers = [_COLS_GASTOS.get(i, "") for i in range(1, n + 1)]
    _titulos(ws, headers, ancho_min=10)
    ejemplo = {
        1: "Arrayanes 40", 2: "2", 3: "EXCAVACIONES, FUNDACIONES Y CONCRETOS",
        4: "Corte 1", 5: "2.02", 6: "Fundaciones", 7: "2025-03-15",
        12: "Ferretería El Roble SAS", 13: "900123456", 15: "Factura de venta",
        16: "FE-4256", 17: "Cemento y varilla", 18: 5000000, 19: 0, 20: 950000,
        26: 5000000, 27: 125000, 29: 5825000, 30: "Crédito", 31: "Pendiente de pago",
        32: "Cuentas x pagar", 33: "Empresa", 34: "", 36: "2025-04-15",
        37: "compras", 41: "", 42: 0, 44: 0, 46: 5825000, 47: "No", 48: "14%",
        49: 815500,
    }
    ws.append([ejemplo.get(i, "") for i in range(1, n + 1)])
    _hoja_instrucciones(wb, [
        ["(importante)", "La hoja se lee por POSICIÓN de columna: no muevas, "
         "insertes ni borres columnas; deja en blanco las que no uses.",
         "El importador cruza contra las facturas que ya llegaron por correo."],
        ["Proyecto / Capítulo / Actividad", "Nombres tal como están en la app; "
         "los ID de capítulo/actividad son opcionales.", "Arrayanes 40 · 2.02"],
        ["Estado", "Estado de pago de la factura.", "Pagada / Pendiente de pago / "
         "Pendiente reporte pago / Parcialmente pagada / Anulada"],
        ["Forma de pago", "", "Contado / Crédito / Abono / Legalización anticipo / Anulada"],
        ["Medio de pago", "", "Cuentas x pagar / Efectivo / Cheque / Tarjeta crédito / ..."],
        ["Pagador", "Quién paga.", "Empresa / Cliente"],
        ["Saldo calculado", "El saldo que ya trae la resta hecha en el Excel.", "5825000"],
        ["Valores", "En pesos, sin puntos ni signo $.", "5825000"],
    ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
